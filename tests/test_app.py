"""
End-to-end test of the Streamlit interface.

Runs app.py exactly as the browser would, with the file uploader replaced by a
real workbook, then clicks "Upload & Audit" and checks the results and the
feedback actually render without errors.

Run from the project folder:   python -m tests.test_app  [path-to-xlsx ...]
"""

from __future__ import annotations

import io
import pathlib
import sys

import streamlit as st
from streamlit.testing.v1 import AppTest

sys.path.insert(0, ".")

PASSED, FAILED = [], []


def expect(name, condition, extra=""):
    (PASSED if condition else FAILED).append(name)
    print(("  PASS  " if condition else "  FAIL  ") + name
          + (f"   {extra}" if extra and not condition else ""))


class FakeUpload(io.BytesIO):
    """Stands in for a Streamlit UploadedFile."""

    def __init__(self, path: pathlib.Path):
        data = path.read_bytes()
        super().__init__(data)
        self.name = path.name
        self._data = data

    def getvalue(self):
        return self._data


def click_button(app, label: str):
    button = next((b for b in app.button if b.label == label), None)
    assert button is not None, f"no button labelled {label!r} (have: {[b.label for b in app.button]})"
    return button.click().run()


def run_app(path: pathlib.Path):
    upload = FakeUpload(path)
    original = st.file_uploader
    st.file_uploader = lambda *a, **k: upload      # noqa: ARG005
    try:
        app = AppTest.from_file("app.py", default_timeout=180)
        app.run()
        expect(f"{path.name}: page renders before the audit", not app.exception,
               str(app.exception))
        expect(f"{path.name}: audit button is present",
               any(b.label == "Upload & Audit" for b in app.button))
        click_button(app, "Upload & Audit")
        return app
    finally:
        st.file_uploader = original


def check(path: pathlib.Path):
    print(f"\n{path.name}")
    app = run_app(path)

    expect(f"{path.name}: audit runs without an exception",
           not app.exception, str(app.exception))
    expect(f"{path.name}: no error message shown",
           len(app.error) == 0, str([e.value for e in app.error]))

    text = " ".join(str(m.value) for m in app.markdown)
    expect(f"{path.name}: results heading rendered", "Issues found" in text
           or any("No data-quality issues" in str(s.value) for s in app.success))
    expect(f"{path.name}: feedback document rendered", "Feedback on" in text,
           text[:200])

    metrics = {m.label: m.value for m in app.metric}
    expect(f"{path.name}: rows-checked metric present",
           "Asset rows checked" in metrics, str(metrics))
    expect(f"{path.name}: findings metric present", "Findings" in metrics, str(metrics))

    codes = " ".join(str(c.value) for c in app.code)
    expect(f"{path.name}: plain-text copy box has the feedback",
           "Feedback on" in codes and "Please:" in codes)

    downloads = [b.label for b in app.get("download_button")]
    for wanted in ["Download .txt", "Download .docx", "Download .pdf",
                   "Detailed findings (.csv)"]:
        expect(f"{path.name}: {wanted} button present", wanted in downloads,
               str(downloads))

    print(f"    rows checked: {metrics.get('Asset rows checked')} | "
          f"sections: {metrics.get('Sections needing correction')} | "
          f"findings: {metrics.get('Findings')}")
    return app


def check_preprocess(path: pathlib.Path):
    print(f"\n[preprocess] {path.name}")
    upload = FakeUpload(path)
    original = st.file_uploader
    st.file_uploader = lambda *a, **k: upload      # noqa: ARG005
    try:
        app = AppTest.from_file("app.py", default_timeout=180)
        app.run()
        expect(f"{path.name}: preprocess button is present",
               any(b.label == "Clean & Prepare Download" for b in app.button))
        click_button(app, "Clean & Prepare Download")
    finally:
        st.file_uploader = original

    expect(f"{path.name}: cleaning runs without an exception",
           not app.exception, str(app.exception))
    expect(f"{path.name}: no error message shown",
           len(app.error) == 0, str([e.value for e in app.error]))

    subheaders = [s.value for s in app.subheader]
    expect(f"{path.name}: cleaning-complete heading rendered",
           "Cleaning complete" in subheaders, str(subheaders))

    metrics = {m.label: m.value for m in app.metric}
    for wanted in ["Sheets cleaned", "Rows processed", "Changes made", "Dates left untouched"]:
        expect(f"{path.name}: {wanted} metric present", wanted in metrics, str(metrics))

    downloads = [b.label for b in app.get("download_button")]
    expect(f"{path.name}: cleaned workbook download present",
           "Download cleaned .xlsx" in downloads, str(downloads))
    expect(f"{path.name}: summary csv download present",
           "Summary (.csv)" in downloads, str(downloads))

    print(f"    sheets cleaned: {metrics.get('Sheets cleaned')} | "
          f"rows processed: {metrics.get('Rows processed')} | "
          f"changes made: {metrics.get('Changes made')} | "
          f"dates untouched: {metrics.get('Dates left untouched')}")

    # Verify the actual bytes behind the download button are a real, readable
    # workbook with every original sheet present — not just that a button
    # with the right label exists.
    cleaned_bytes = app.session_state["preprocess_book_bytes"]
    expect(f"{path.name}: the cleaned file is non-trivial in size", len(cleaned_bytes) > 1000,
           f"{len(cleaned_bytes)} bytes")

    import io as _io
    import openpyxl as _openpyxl
    workbook = _openpyxl.load_workbook(_io.BytesIO(cleaned_bytes))
    expect(f"{path.name}: the cleaned file still has an Asset | GPE Information sheet",
           "Asset | GPE Information" in workbook.sheetnames, str(workbook.sheetnames))
    expect(f"{path.name}: sheets outside the three preprocessed ones are still present "
           f"(nothing else was dropped from the workbook)",
           "Instructions" in workbook.sheetnames, str(workbook.sheetnames))
    return app


if __name__ == "__main__":
    given = [pathlib.Path(a) for a in sys.argv[1:]]
    downloads = pathlib.Path.home() / "Downloads"
    default = [
        downloads / "Data Upload Template v3_Malawi_Zimbabwe.xlsx",
        downloads / "Data Upload Template V3_Zambia RM.xlsx",
    ]
    paths = given or [p for p in default if p.exists()]

    if not paths:
        print("No workbook found to test with.")
        sys.exit(1)

    for path in paths:
        check(path)
        check_preprocess(path)

    print(f"\n{len(PASSED)} passed, {len(FAILED)} failed")
    if FAILED:
        for name in FAILED:
            print("  FAILED:", name)
        sys.exit(1)
