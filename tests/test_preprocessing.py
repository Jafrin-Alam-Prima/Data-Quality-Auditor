"""
Tests for the data-cleaning (preprocessing) feature.

Builds small workbooks in memory, runs clean_workbook() on them, then
re-opens the SAVED output to check the actual cells it wrote -- not just the
in-memory report -- since a migration file is only as good as what actually
lands in the saved .xlsx.

Run from the project folder:   python -m tests.test_preprocessing
"""

from __future__ import annotations

import datetime as _dt
import io
import sys

import openpyxl

sys.path.insert(0, ".")

from auditor.dates import parse_date            # noqa: E402
from auditor.preprocessing import clean_workbook  # noqa: E402

PASSED, FAILED = [], []


def expect(name, condition, extra=""):
    (PASSED if condition else FAILED).append(name)
    print(("  PASS  " if condition else "  FAIL  ") + name + (f"   {extra}" if extra and not condition else ""))


def build_workbook(supplier_rows=None, office_rows=None, asset_rows=None,
                   include_instructions=True, skip_supplier=False, skip_office=False):
    book = openpyxl.Workbook()
    book.remove(book.active)

    if include_instructions:
        sheet = book.create_sheet("Instructions")
        sheet.append(["Please fill in the template carefully."])
        sheet.append(["Row 2 of the instructions, unrelated to any data sheet."])

    if not skip_supplier:
        sheet = book.create_sheet("Supplier")
        sheet.append(["SupplierName", "Code", "Country", "Region", "Email", "Status"])
        for row in (supplier_rows or []):
            sheet.append(row)

    if not skip_office:
        sheet = book.create_sheet("Office")
        sheet.append(["Region", "Country", "OfficeName", "OfficeType", "DataCollectionDate"])
        for row in (office_rows or []):
            sheet.append(row)

    sheet = book.create_sheet("Asset | GPE Information")
    sheet.append([None] * 6)
    sheet.append(["Asset Name", "Asset ID", "Purchase Date", "Warranty End Date",
                  "Assigned Date", "Supplier"])
    for row in (asset_rows or []):
        sheet.append(row)

    buffer = io.BytesIO()
    book.save(buffer)
    buffer.seek(0)
    return buffer


def clean_and_reopen(day_first=True, **kwargs):
    source = build_workbook(**kwargs)
    book, result = clean_workbook(source, day_first=day_first, filename="Testland.xlsx")
    out = io.BytesIO()
    book.save(out)
    out.seek(0)
    reopened = openpyxl.load_workbook(out, data_only=True)
    return result, reopened


def summary(result, key):
    return next(s for s in result.summaries if s.key == key)


# ---------------------------------------------------------------------------

def test_supplier_blank_and_placeholder_removed():
    result, book = clean_and_reopen(supplier_rows=[
        ["Alpha Traders", "S1", "Zambia", "Lusaka", "a@x.com", "Active"],
        ["", "S2", "Zambia", "Lusaka", "b@x.com", "Active"],
        ["N/A", "S3", "Zambia", "Lusaka", "c@x.com", "Active"],
        ["  ", "S4", "Zambia", "Lusaka", "d@x.com", "Active"],
    ])
    s = summary(result, "supplier_ref")
    expect("blank, N/A and whitespace-only supplier names are all removed",
           s.removed_blank_or_placeholder == 3, str(s))
    expect("the one real supplier survives", s.remaining_rows == 1, str(s))

    names = [row[0].value for row in book["Supplier"].iter_rows(min_row=2) if row[0].value]
    expect("the saved file actually only has the real supplier left",
           names == ["Alpha Traders"], str(names))


def test_supplier_duplicate_keeps_most_complete():
    result, book = clean_and_reopen(supplier_rows=[
        ["Capital Computers", None, None, None, None, None],
        ["capital computers", "S9", "Zambia", "Lusaka", "info@cc.com", "Active"],
        ["CAPITAL COMPUTERS", "S9", None, None, None, "Active"],
        ["Beta Motors", "S2", "Zambia", "Ndola", "b@x.com", "Active"],
    ])
    s = summary(result, "supplier_ref")
    expect("two of the three duplicate spellings are removed", s.removed_duplicates == 2, str(s))
    expect("the surviving row is the most complete one",
           s.duplicate_examples == ["capital computers"], str(s.duplicate_examples))

    rows = [tuple(c.value for c in row) for row in book["Supplier"].iter_rows(min_row=2)
           if row[0].value]
    names = [r[0] for r in rows]
    expect("both distinct suppliers survive, no more", sorted(n.lower() for n in names)
           == ["beta motors", "capital computers"], str(names))
    kept = next(r for r in rows if r[0].lower() == "capital computers")
    expect("the kept row is the fully filled-in one (has Code, Country, Email)",
           kept[1] == "S9" and kept[4] == "info@cc.com", str(kept))


def test_office_duplicate_and_blank_name_and_date():
    result, book = clean_and_reopen(office_rows=[
        ["ESA", "Zambia", "Lusaka", "Country Office", "26/12/2018"],
        ["ESA", "Zambia", "lusaka", None, None],
        ["ESA", "Zambia", "", "Field Office", "01/01/2020"],   # blank name, kept untouched
        ["ESA", "Zambia", "Kitwe", "Field Office", _dt.datetime(2019, 5, 1)],
    ])
    s = summary(result, "office_ref")
    expect("the duplicate office name is removed, keeping the more complete row",
           s.removed_duplicates == 1, str(s))
    expect("blank office name row is not touched by dedup, still present",
           s.remaining_rows == 3, str(s))
    expect("the DataCollectionDate column is detected as a date column",
           s.date_columns == ["DataCollectionDate"], str(s.date_columns))
    expect("two real dates reformatted (blank office row's date is untouched "
           "since 01/01/2020 there is still a parseable value, so 3 total)",
           s.dates_reformatted == 3, str(s))

    rows = [tuple(c.value for c in row) for row in book["Office"].iter_rows(min_row=2)]
    lusaka = next(r for r in rows if str(r[2]).lower() == "lusaka")
    expect("Lusaka's surviving row kept the real date, not the blank duplicate",
           isinstance(lusaka[4], _dt.datetime) and lusaka[4] == _dt.datetime(2018, 12, 26),
           str(lusaka))
    kitwe = next(r for r in rows if r[2] == "Kitwe")
    expect("a cell that was already a real datetime object is preserved correctly",
           kitwe[4] == _dt.datetime(2019, 5, 1), str(kitwe))


def test_asset_info_dates_reformatted_and_unparseable_reported():
    result, book = clean_and_reopen(asset_rows=[
        ["Chair", "A1", "26/12/2018", "01/01/2022", _dt.datetime(2020, 3, 4), "Alpha"],
        ["Desk", "A2", "11", "2021-07-15", "not a date", "Alpha"],   # bad Purchase Date
        ["Lamp", "A3", "", "", "", "Alpha"],                        # blank dates: not reported
    ])
    s = summary(result, "asset_info")
    expect("all three date columns detected",
           set(s.date_columns) == {"Purchase Date", "Warranty End Date", "Assigned Date"},
           str(s.date_columns))
    expect("5 real dates reformatted (row1 x3, row2's ISO Warranty End Date, "
           "blank cells don't count)",
           s.dates_reformatted == 4, str(s))
    expect("exactly 2 unparseable dates reported (bare '11', 'not a date')",
           len(s.dates_unparseable) == 2, str(s.dates_unparseable))
    bad_values = {u.value for u in s.dates_unparseable}
    expect("the unparseable list names the exact bad values",
           bad_values == {"11", "not a date"}, str(bad_values))
    expect("the unparseable list gives the real Excel row for each",
           {u.row for u in s.dates_unparseable} == {4}, str(s.dates_unparseable))

    rows = [tuple(c.value for c in row) for row in book["Asset | GPE Information"].iter_rows(min_row=3)]
    row1 = rows[0]
    expect("row 1's Purchase Date became a real datetime, correctly parsed as day-first",
           row1[2] == _dt.datetime(2018, 12, 26), str(row1))
    row2 = rows[1]
    expect("row 2's unparseable Purchase Date ('11') is left exactly as it was",
           row2[2] == "11", str(row2))
    expect("row 2's valid ISO Warranty End Date was still reformatted",
           row2[3] == _dt.datetime(2021, 7, 15), str(row2))


def test_day_first_vs_month_first():
    expect("13/02/2020 with day_first is 13 Feb (unambiguous either way)",
           parse_date("13/02/2020", day_first=True) == _dt.datetime(2020, 2, 13))
    expect("02/13/2020 with day_first=False is 13 Feb",
           parse_date("02/13/2020", day_first=False) == _dt.datetime(2020, 2, 13))
    expect("01/02/2020 with day_first=True is 1 Feb",
           parse_date("01/02/2020", day_first=True) == _dt.datetime(2020, 2, 1))
    expect("01/02/2020 with day_first=False is 2 Jan",
           parse_date("01/02/2020", day_first=False) == _dt.datetime(2020, 1, 2))
    expect("an ISO date ignores day_first entirely",
           parse_date("2020-02-01", day_first=False) == _dt.datetime(2020, 2, 1))
    expect("a plainly invalid date (32nd of a month) is rejected",
           parse_date("32/01/2020", day_first=True) is None)
    expect("a bare small number outside any plausible date range is rejected, "
           "not silently turned into an early-1900 Excel serial date",
           parse_date(3, day_first=True) is None)
    expect("the exact real-world corrupted value found in production ('11', as "
           "a leftover fragment, not a real date) is rejected the same way",
           parse_date("11", day_first=True) is None and parse_date(11, day_first=True) is None)
    expect("the same corrupted value as '12' is also rejected",
           parse_date("12", day_first=True) is None)
    expect("a plausible Excel serial date number is accepted",
           parse_date(44197, day_first=True) == _dt.datetime(2021, 1, 1))
    expect("None is not an error, just \"nothing to parse\"",
           parse_date(None, day_first=True) is None)


def test_other_sheets_are_never_touched():
    result, book = clean_and_reopen(
        supplier_rows=[["Alpha Traders", "S1", "Zambia", "Lusaka", "a@x.com", "Active"]],
    )
    instructions = [c.value for row in book["Instructions"].iter_rows() for c in row if c.value]
    expect("a sheet outside the three preprocessed ones is completely unchanged",
           instructions == ["Please fill in the template carefully.",
                            "Row 2 of the instructions, unrelated to any data sheet."],
           str(instructions))


def test_missing_sheet_is_reported_but_others_still_run():
    result, book = clean_and_reopen(
        skip_supplier=True,
        office_rows=[["ESA", "Zambia", "Lusaka", "Country Office", "26/12/2018"]],
    )
    expect("the missing Supplier sheet is reported by name",
           "Supplier" in result.missing_sheets, str(result.missing_sheets))
    expect("the Office sheet was still cleaned despite Supplier being missing",
           any(s.key == "office_ref" for s in result.summaries))


def test_source_file_on_disk_is_never_modified():
    import os
    import tempfile

    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    try:
        source = build_workbook(supplier_rows=[
            ["Alpha Traders", "S1", "Zambia", "Lusaka", "a@x.com", "Active"],
            ["", "S2", "Zambia", "Lusaka", "b@x.com", "Active"],
        ])
        with open(path, "wb") as f:
            f.write(source.getvalue())
        original_bytes = open(path, "rb").read()

        clean_workbook(path, day_first=True, filename="Testland.xlsx")

        after_bytes = open(path, "rb").read()
        expect("the uploaded file on disk is byte-for-byte unchanged after cleaning",
               original_bytes == after_bytes)
    finally:
        os.remove(path)


if __name__ == "__main__":
    for test in [test_supplier_blank_and_placeholder_removed,
                test_supplier_duplicate_keeps_most_complete,
                test_office_duplicate_and_blank_name_and_date,
                test_asset_info_dates_reformatted_and_unparseable_reported,
                test_day_first_vs_month_first,
                test_other_sheets_are_never_touched,
                test_missing_sheet_is_reported_but_others_still_run,
                test_source_file_on_disk_is_never_modified]:
        print(f"\n{test.__name__}")
        test()

    print(f"\n{len(PASSED)} passed, {len(FAILED)} failed")
    if FAILED:
        for name in FAILED:
            print("  FAILED:", name)
        sys.exit(1)
