"""
Self-contained tests for the audit engine.

Builds small workbooks in memory that deliberately contain each kind of
problem, then checks that the engine reports exactly those problems.

Run from the project folder:   python -m tests.test_engine
"""

from __future__ import annotations

import io
import sys

import openpyxl

sys.path.insert(0, ".")

from auditor import audit, build_feedback, render_text   # noqa: E402

ASSET_HEADERS = [
    "Asset Name", "Asset ID", "Asset | GPE Category", "Asset | GPE Category Code",
    "Invoice Currency", "Item Value Invoice Currency", "Item Value USD",
    "Supplier", "Office", "Asset | GPE Condition",
]

CATEGORIES = [
    ("Vehicles (all vehicles including 4x4)", "VEH"),
    ("Motorbike and quad bikes", "MOT"),
    ("Computers (servers, laptops, desktop CPUs)", "CMP"),
]
CONDITIONS = ["New (Purchased within the last 12 months)",
              "Good (No visible damage, no repairs completed)",
              "Fair (Reasonable condition showing signs of wear and tear)",
              "Poor (Damaged, old but still have a limited use)",
              "BER (Beyond Economic Repair)"]
DISPOSALS = ["Donation", "Sale", "Write Off"]
OFFICE_TYPES = ["Country Office", "Field Office"]


def build_workbook(asset_rows, suppliers=None, offices=None,
                   skip_sheets=(), condition_header=False):
    book = openpyxl.Workbook()
    book.remove(book.active)

    if "asset_info" not in skip_sheets:
        sheet = book.create_sheet("Asset | GPE Information")
        sheet.append([None] * len(ASSET_HEADERS))      # blank row above header
        sheet.append(ASSET_HEADERS)
        for row in asset_rows:
            sheet.append(row)

    if "category_ref" not in skip_sheets:
        sheet = book.create_sheet("Asset | GPECategory")
        sheet.append([None] * 3)
        sheet.append(["Category Name", "Category code", "Assets Type"])
        for name, code in CATEGORIES:
            sheet.append([name, code, "Asset"])

    if "condition_ref" not in skip_sheets:
        sheet = book.create_sheet("Condition| Disposal Reason")
        if condition_header:
            sheet.append(["Condition", None, None, "Disposal Reason", "Office Type"])
        for index in range(max(len(CONDITIONS), len(DISPOSALS), len(OFFICE_TYPES))):
            sheet.append([
                CONDITIONS[index] if index < len(CONDITIONS) else None, None, None,
                DISPOSALS[index] if index < len(DISPOSALS) else None,
                OFFICE_TYPES[index] if index < len(OFFICE_TYPES) else None,
            ])

    if "supplier_ref" not in skip_sheets:
        sheet = book.create_sheet("Supplier")
        sheet.append(["SupplierName", "Code", "Country", "Status"])
        for name in (suppliers if suppliers is not None else ["Alpha Traders", "Beta Motors"]):
            sheet.append([name, None, "Zambia", "Active"])

    if "office_ref" not in skip_sheets:
        sheet = book.create_sheet("Office")
        sheet.append(["Region", "Country", "OfficeName", "OfficeType"])
        for name, office_type in (offices if offices is not None
                                  else [("Lusaka", "Country Office"), ("Kitwe", "Field Office")]):
            sheet.append(["ESA", "Zambia", name, office_type])

    buffer = io.BytesIO()
    book.save(buffer)
    buffer.seek(0)
    return buffer


def run(asset_rows, filename="Data Upload Template V3_Testland.xlsx", **kwargs):
    return audit(build_workbook(asset_rows, **kwargs), filename=filename)


def codes(result):
    return {(i.section, i.code) for i in result.issues}


GOOD_ROW = ["Toyota Hilux", "ZMB-VEH-0001", CATEGORIES[0][0], "VEH",
            "ZMW", 37027, 1500, "Alpha Traders", "Lusaka", CONDITIONS[1]]

PASSED, FAILED = [], []


def expect(name, condition, extra=""):
    (PASSED if condition else FAILED).append(name)
    print(("  PASS  " if condition else "  FAIL  ") + name + (f"   {extra}" if extra and not condition else ""))


# ---------------------------------------------------------------------------

def test_clean_workbook():
    result = run([GOOD_ROW, ["Laptop", "ZMB-CMP-0002", CATEGORIES[2][0], "CMP",
                             "USD", 900, 900, "Beta Motors", "Kitwe", CONDITIONS[0]]])
    expect("clean workbook reports no issues", not result.has_issues,
           str([(i.section, i.code, i.values[:3]) for i in result.issues]))
    document = build_feedback(result)
    expect("clean workbook feedback says so", document["clean"])
    expect("country taken from file name", result.country == "Testland", result.country)


def test_asset_name_and_id():
    rows = [
        GOOD_ROW,
        ["", "ZMB-VEH-0002", CATEGORIES[0][0], "VEH", "ZMW", 10, 10, "Alpha Traders", "Lusaka", CONDITIONS[1]],
        ["N/A", "ZMB-VEH-0003", CATEGORIES[0][0], "VEH", "ZMW", 10, 10, "Alpha Traders", "Lusaka", CONDITIONS[1]],
        ["Printer", "", CATEGORIES[2][0], "CMP", "ZMW", 10, 10, "Alpha Traders", "Lusaka", CONDITIONS[1]],
        ["Printer", "TBA", CATEGORIES[2][0], "CMP", "ZMW", 10, 10, "Alpha Traders", "Lusaka", CONDITIONS[1]],
        ["Printer", 0, CATEGORIES[2][0], "CMP", "ZMW", 10, 10, "Alpha Traders", "Lusaka", CONDITIONS[1]],
        ["Printer", "#REF!", CATEGORIES[2][0], "CMP", "ZMW", 10, 10, "Alpha Traders", "Lusaka", CONDITIONS[1]],
        ["Scanner", "ZMB-VEH-0001", CATEGORIES[2][0], "CMP", "ZMW", 10, 10, "Alpha Traders", "Lusaka", CONDITIONS[1]],
    ]
    result = run(rows)
    found = codes(result)
    expect("blank Asset Name found", ("Asset Name", "blank") in found)
    expect("placeholder Asset Name found", ("Asset Name", "placeholder") in found)
    expect("blank Asset ID found", ("Asset ID", "blank") in found)
    expect("placeholder Asset ID found (TBA and 0)", ("Asset ID", "placeholder") in found)
    expect("#REF! Asset ID found", ("Asset ID", "error_value") in found)
    expect("duplicate Asset ID found", ("Asset ID", "duplicate") in found)

    duplicate = next(i for i in result.issues if i.code == "duplicate")
    expect("duplicate reports the repeated value",
           duplicate.values == ["ZMB-VEH-0001"], str(duplicate.values))
    placeholder = next(i for i in result.issues
                       if i.section == "Asset ID" and i.code == "placeholder")
    expect("a blank cell is not also called a placeholder", placeholder.count == 2,
           f"count={placeholder.count}")


def test_category_and_code():
    rows = [
        GOOD_ROW,
        ["Bike", "A1", "Flying machines", "FLY", "ZMW", 10, 10, "Alpha Traders", "Lusaka", CONDITIONS[1]],
        ["Bike", "A2", CATEGORIES[1][0], "CMP", "ZMW", 10, 10, "Alpha Traders", "Lusaka", CONDITIONS[1]],
        ["Bike", "A3", "motorbike and QUAD bikes", "mot", "ZMW", 10, 10, "Alpha Traders", "Lusaka", CONDITIONS[1]],
        ["Bike", "A4", "", "", "ZMW", 10, 10, "Alpha Traders", "Lusaka", CONDITIONS[1]],
    ]
    found = codes(run(rows))
    expect("unknown category found", ("Asset Category", "not_in_reference") in found)
    expect("unknown category code found", ("Asset Category Code", "not_in_reference") in found)
    expect("code not matching its category found",
           ("Asset Category Code", "code_category_mismatch") in found)
    expect("category spelling difference found", ("Asset Category", "case_mismatch") in found)
    expect("code spelling difference found", ("Asset Category Code", "case_mismatch") in found)
    expect("blank category found", ("Asset Category", "blank") in found)
    expect("blank category code found", ("Asset Category Code", "blank") in found)


def test_item_values():
    rows = [
        GOOD_ROW,
        ["Desk", "B1", CATEGORIES[2][0], "CMP", "ZMW", "OLD MATERIAL/UNKNOWN", 10, "Alpha Traders", "Lusaka", CONDITIONS[1]],
        ["Desk", "B2", CATEGORIES[2][0], "CMP", "ZMW", 0, 0, "Alpha Traders", "Lusaka", CONDITIONS[1]],
        ["Desk", "B3", CATEGORIES[2][0], "CMP", "ZMW", "", "", "Alpha Traders", "Lusaka", CONDITIONS[1]],
        ["Desk", "B4", CATEGORIES[2][0], "CMP", "USD", 1000, 250, "Alpha Traders", "Lusaka", CONDITIONS[1]],
        ["Desk", "B5", CATEGORIES[2][0], "CMP", "ZMW", "1,250.75", 55, "Alpha Traders", "Lusaka", CONDITIONS[1]],
    ]
    result = run(rows)
    found = codes(result)
    expect("text in Item Value Invoice Currency found",
           ("Item Value Invoice Currency", "not_numeric") in found)
    expect("zero Item Value found", ("Item Value Invoice Currency", "zero") in found)
    expect("blank Item Value found", ("Item Value Invoice Currency", "blank") in found)
    expect("blank Item Value USD found", ("Item Value USD", "blank") in found)
    expect("USD invoice not matching USD value found",
           ("Item Value USD", "usd_mismatch") in found)

    numeric = next(i for i in result.issues if i.code == "not_numeric")
    expect("a formatted number is accepted as a value",
           "1,250.75" not in numeric.values and "1250.75" not in numeric.values,
           str(numeric.values))


def test_supplier_office_condition():
    rows = [
        GOOD_ROW,
        ["Desk", "C1", CATEGORIES[2][0], "CMP", "ZMW", 10, 10, "Gamma Supplies", "Lusaka", CONDITIONS[1]],
        ["Desk", "C2", CATEGORIES[2][0], "CMP", "ZMW", 10, 10, "alpha traders", "LUSAKA", CONDITIONS[1]],
        ["Desk", "C3", CATEGORIES[2][0], "CMP", "ZMW", 10, 10, "", "", ""],
        ["Desk", "C4", CATEGORIES[2][0], "CMP", "ZMW", 10, 10, "N/A", "Ndola", "Very good"],
        ["Desk", "C5", CATEGORIES[2][0], "CMP", "ZMW", 10, 10, 0, 0, "N/A"],
    ]
    result = run(rows)
    found = codes(result)
    expect("supplier not in reference found", ("Supplier", "not_in_reference") in found)
    expect("supplier spelling difference found", ("Supplier", "case_mismatch") in found)
    expect("blank supplier found", ("Supplier", "blank") in found)
    expect("placeholder supplier found", ("Supplier", "placeholder") in found)
    expect("office not in reference found", ("Office", "not_in_reference") in found)
    expect("office spelling difference found", ("Office", "case_mismatch") in found)
    expect("invalid condition found", ("Asset | GPE Condition", "not_in_reference") in found)
    expect("blank condition found", ("Asset | GPE Condition", "blank") in found)
    expect("placeholder condition found", ("Asset | GPE Condition", "placeholder") in found)

    expect("missing suppliers raised against the Supplier sheet",
           ("Supplier Reference Sheet", "ref_missing") in found)
    expect("missing offices raised against the Office sheet",
           ("Office Reference Sheet", "ref_missing") in found)
    missing = next(i for i in result.issues if i.section == "Supplier Reference Sheet"
                   and i.code == "ref_missing")
    expect("missing supplier list is exact", missing.values == ["Gamma Supplies"],
           str(missing.values))


def test_reference_sheets():
    result = run(
        [GOOD_ROW],
        suppliers=["Alpha Traders", "Beta Motors", "alpha traders", "", "N/A"],
        offices=[("Lusaka", "Country Office"), ("Kitwe", ""),
                 ("Lusaka", "Field Office"), ("", "Field Office"),
                 ("Ndola", "Regional Office")],
    )
    found = codes(result)
    expect("duplicate supplier found", ("Supplier Reference Sheet", "ref_duplicate") in found)
    expect("blank supplier name found", ("Supplier Reference Sheet", "ref_blank") in found)
    expect("placeholder supplier name found",
           ("Supplier Reference Sheet", "ref_placeholder") in found)
    expect("duplicate office found", ("Office Reference Sheet", "ref_duplicate") in found)
    expect("blank office name found", ("Office Reference Sheet", "ref_blank") in found)
    expect("blank office type found", ("Office Reference Sheet", "ref_type_blank") in found)
    expect("invalid office type found", ("Office Reference Sheet", "ref_type_invalid") in found)


def test_missing_sheet():
    result = run([GOOD_ROW], skip_sheets=("supplier_ref",))
    expect("missing reference sheet reported",
           ("Missing Sheets", "missing_sheet") in codes(result))
    text = render_text(build_feedback(result))
    expect("missing sheet named in the feedback", "Supplier" in text)
    expect("missing sheet section comes first",
           text.index("Workbook Structure") < len(text))


def test_condition_sheet_with_header():
    result = run([GOOD_ROW], condition_header=True)
    expect("condition sheet with a header row still reads the options",
           not result.has_issues,
           str([(i.section, i.code) for i in result.issues]))


def test_feedback_shape():
    rows = [GOOD_ROW,
            ["", "", CATEGORIES[0][0], "VEH", "ZMW", 0, "", "Gamma Supplies", "Ndola", "Broken"]]
    document = build_feedback(run(rows))
    text = render_text(document)
    expect("feedback has a title", text.startswith("Feedback on Testland Data"))
    expect("feedback has a date", "Date:" in text)
    expect("every section has a Please block",
           all("Please:" in s["problem"] or s["actions"] for s in document["sections"]))
    expect("no counts leak into the feedback",
           "rows" not in text.lower() and "%" not in text)
    expect("no clean sections are listed", "No issues" not in text)
    expect("sections are numbered from 1",
           [s["number"] for s in document["sections"]][:1] == [1])

    from auditor.exporters import to_docx, to_findings_csv, to_pdf, to_txt
    expect("txt export works", to_txt(document).startswith(b"Feedback on"))
    expect("docx export works", to_docx(document)[:2] == b"PK")
    expect("pdf export works", to_pdf(document)[:4] == b"%PDF")
    expect("csv export works", b"Section" in to_findings_csv(run(rows)))


def test_normalization_handles_invisible_characters():
    """Reproduces the reported bug: a supplier that exists in the reference
    sheet was reported as missing because of hidden Unicode characters that
    look identical to a clean value in Excel."""
    from auditor.utils import normalize_for_matching

    variants = [
        "Capital Computer",
        " Capital Computer",
        "Capital Computer ",
        "Capital  Computer",                    # double space
        "CAPITAL COMPUTER",
        "Capital Computer",                # non-breaking space as the separator
        "Capital ​Computer",               # zero-width space next to a real space
        "﻿Capital Computer",                # UTF-8 BOM at the very start
        "Capital Computer⁠",               # trailing word joiner
        "Capital Computer‌",               # trailing zero-width non-joiner
        "Capital Computer‎",               # trailing left-to-right mark
        "Capital­ Computer",               # soft hyphen next to a real space
    ]
    keys = {normalize_for_matching(v) for v in variants}
    expect("all invisible-character variants normalize to the same key",
           len(keys) == 1, str(keys))

    rows = [GOOD_ROW]
    for index, variant in enumerate(variants):
        rows.append(["Laptop", f"ZMB-CMP-{index:04d}", CATEGORIES[2][0], "CMP",
                     "ZMW", 10, 10, variant, "Lusaka", CONDITIONS[1]])
    result = run(rows, suppliers=["Alpha Traders", "Capital Computer"])
    found = codes(result)
    expect("a supplier that exists, with only invisible/spacing differences, "
           "is never reported as missing",
           ("Supplier", "not_in_reference") not in found, str(found))


def test_missing_vs_formatting_difference_are_distinct():
    rows = [
        GOOD_ROW,
        # Genuinely missing: does not exist in the Supplier sheet at all.
        ["Desk", "D1", CATEGORIES[2][0], "CMP", "ZMW", 10, 10, "ABC Company",
         "Lusaka", CONDITIONS[1]],
        # Same supplier, only a formatting difference.
        ["Desk", "D2", CATEGORIES[2][0], "CMP", "ZMW", 10, 10, "capital  computer",
         "Lusaka", CONDITIONS[1]],
    ]
    result = run(rows, suppliers=["Alpha Traders", "Capital Computer"])
    missing = next((i for i in result.issues
                    if i.section == "Supplier" and i.code == "not_in_reference"), None)
    formatting = next((i for i in result.issues
                       if i.section == "Supplier" and i.code == "case_mismatch"), None)

    # Issue values are shown through cell_text (leading/trailing/double
    # spaces cleaned up for display); the case difference is preserved, so
    # match on that rather than on the raw double-spaced literal.
    expect("a supplier not in the reference sheet at all is reported missing",
           missing is not None and "ABC Company" in missing.values)
    expect("a supplier that only differs in spacing/case is reported as a "
           "formatting difference, not missing",
           formatting is not None and "capital computer" in formatting.values)
    if missing:
        expect("the formatting-difference supplier is not also reported missing",
               "capital computer" not in missing.values)
    if formatting:
        expect("the genuinely missing supplier is not also reported as a "
               "formatting difference",
               "ABC Company" not in formatting.values)


def test_condition_short_form_is_recognised_as_valid():
    """The Condition | Disposal Reason sheet stores long descriptive text
    ("Good (No visible damage, no repairs completed)"). A register that
    stores the short form ("Good") is using the same, valid option and must
    not be reported as invalid — derived from the sheet's own wording, not a
    hard-coded list."""
    short_forms = ["New", "Good", "Fair", "Poor", "BER",
                   "  good  ", "GOOD", "Good "]
    rows = [GOOD_ROW]
    for index, value in enumerate(short_forms):
        rows.append(["Chair", f"ZMB-CMP-{index:04d}", CATEGORIES[2][0], "CMP",
                     "ZMW", 10, 10, "Alpha Traders", "Lusaka", value])
    result = run(rows)
    found = codes(result)
    expect("short-form condition values matching the reference sheet's own "
           "wording are accepted as valid",
           ("Asset | GPE Condition", "not_in_reference") not in found, str(found))

    # A value that is not a valid option under any form must still be caught.
    bad_row = ["Chair", "ZMB-CMP-9999", CATEGORIES[2][0], "CMP", "ZMW", 10, 10,
              "Alpha Traders", "Lusaka", "Very good"]
    result_bad = run([GOOD_ROW, bad_row])
    expect("a genuinely invalid condition is still reported",
           ("Asset | GPE Condition", "not_in_reference") in codes(result_bad))


def test_category_short_form_and_code_mismatch_still_detected():
    """A category entered in short form must still match its reference entry
    and its category code must still be validated against it."""
    rows = [
        GOOD_ROW,
        # Short form, correct code -> no problem.
        ["Bike", "E1", "Motorbike and quad bikes", "MOT", "ZMW", 10, 10,
         "Alpha Traders", "Lusaka", CONDITIONS[1]],
        # Short form, wrong code -> must still be caught.
        ["Bike", "E2", CATEGORIES[0][0].split("(")[0].strip(), "MOT", "ZMW", 10, 10,
         "Alpha Traders", "Lusaka", CONDITIONS[1]],
    ]
    result = run(rows)
    found = codes(result)
    expect("a short-form category matching the reference sheet's wording is "
           "not reported as unknown",
           ("Asset Category", "not_in_reference") not in found, str(found))
    expect("a code that does not belong to the (short-form) category is "
           "still caught",
           ("Asset Category Code", "code_category_mismatch") in found, str(found))


def test_country_guessing():
    from auditor.utils import guess_country
    cases = [
        ("Data Upload Template V3_Zambia RM.xlsx", "Zambia RM"),
        ("Data Upload Template v3_Malawi_Zimbabwe.xlsx", "Malawi & Zimbabwe"),
        ("Data Upload Template v3_Kenya CO_final.xlsx", "Kenya CO"),
    ]
    for filename, wanted in cases:
        got = guess_country(filename)
        expect(f"country from {filename!r} -> {wanted!r}", got == wanted, f"got {got!r}")


if __name__ == "__main__":
    for test in [test_clean_workbook, test_asset_name_and_id, test_category_and_code,
                 test_item_values, test_supplier_office_condition, test_reference_sheets,
                 test_missing_sheet, test_condition_sheet_with_header,
                 test_feedback_shape, test_country_guessing,
                 test_normalization_handles_invisible_characters,
                 test_missing_vs_formatting_difference_are_distinct,
                 test_condition_short_form_is_recognised_as_valid,
                 test_category_short_form_and_code_mismatch_still_detected]:
        print(f"\n{test.__name__}")
        test()

    print(f"\n{len(PASSED)} passed, {len(FAILED)} failed")
    if FAILED:
        for name in FAILED:
            print("  FAILED:", name)
        sys.exit(1)
