"""
Excel processing layer.

Responsibilities:
  * open the workbook safely (read only, never written back)
  * work out which sheet plays which role, even when the names differ
  * find the header row, which is not always row 1
  * expose rows as plain dictionaries with the real Excel row numbers attached

Nothing in this module knows any audit rule.
"""

from __future__ import annotations

import warnings
from dataclasses import dataclass, field
from typing import Any

import openpyxl

from .config import SHEET_SPECS
from .utils import cell_text, is_blank, reduce_name

warnings.filterwarnings("ignore", module="openpyxl")


@dataclass
class Sheet:
    """One worksheet loaded into memory."""

    name: str
    header_row: int                      # 1-based Excel row of the header
    columns: list[str]                   # header labels, in order
    rows: list[dict[str, Any]] = field(default_factory=list)   # label -> value
    row_numbers: list[int] = field(default_factory=list)       # Excel row of each row
    grid: list[list[Any]] = field(default_factory=list)        # raw cells, headers included

    def __len__(self) -> int:
        return len(self.rows)

    def find_column(self, candidates) -> str | None:
        """Return the real header label matching any of the candidate names."""
        wanted = [reduce_name(c) for c in candidates]
        reduced = {reduce_name(c): c for c in self.columns if cell_text(c)}
        for want in wanted:
            if want in reduced:
                return reduced[want]
        # Fall back to a containment match ("Supplier" -> "Supplier Name").
        for want in wanted:
            for key, label in reduced.items():
                if want and (want in key or key in want):
                    return label
        return None

    def column_values(self, label):
        """Yield (excel_row, value) for one column."""
        for number, row in zip(self.row_numbers, self.rows):
            yield number, row.get(label)


@dataclass
class Workbook:
    filename: str
    sheet_names: list[str]
    sheets: dict[str, Sheet]             # logical key -> Sheet
    resolved: dict[str, str]             # logical key -> real sheet name
    missing: list[str]                   # labels of required sheets not found

    def get(self, key: str) -> Sheet | None:
        return self.sheets.get(key)


# ---------------------------------------------------------------------------
# Sheet role detection
# ---------------------------------------------------------------------------

def _score_sheet(spec: dict, sheet_name: str) -> int:
    reduced = reduce_name(sheet_name)
    if not reduced:
        return 0
    if any(f in reduced for f in spec["forbid"]):
        return 0
    if reduced in spec["exact"]:
        return 100
    if not all(f in reduced for f in spec["require_all"]):
        return 0
    if spec["any_of"] and not any(f in reduced for f in spec["any_of"]):
        return 0
    # Shorter names are more likely to be the plain reference sheet.
    return 50 + max(0, 20 - len(reduced))


def detect_sheets(sheet_names: list[str]) -> tuple[dict[str, str], list[str]]:
    """Map each logical sheet key to a real sheet name."""
    resolved: dict[str, str] = {}
    used: set[str] = set()
    missing: list[str] = []

    for spec in SHEET_SPECS:
        best_name, best_score = None, 0
        for name in sheet_names:
            if name in used:
                continue
            score = _score_sheet(spec, name)
            if score > best_score:
                best_name, best_score = name, score
        if best_name:
            resolved[spec["key"]] = best_name
            used.add(best_name)
        elif spec["required"]:
            missing.append(spec["label"])
    return resolved, missing


# ---------------------------------------------------------------------------
# Header row detection
# ---------------------------------------------------------------------------

def _looks_like_header(cells) -> int:
    """Score a row on how much it looks like a row of column titles."""
    score = 0
    for cell in cells:
        text = cell_text(cell)
        if not text:
            continue
        if len(text) > 60:            # long sentences are data, not titles
            score -= 1
            continue
        if isinstance(cell, (int, float)):
            score -= 1
            continue
        score += 1
    return score


def _find_header_row(grid, scan_rows: int = 10) -> int:
    """Return the 0-based index of the most likely header row."""
    best_index, best_score = 0, -10_000
    for index in range(min(scan_rows, len(grid))):
        row = grid[index]
        filled = sum(1 for c in row if not is_blank(c))
        if filled < 2:
            continue
        # A header row must be followed by at least one row with data.
        following = any(
            any(not is_blank(c) for c in grid[j])
            for j in range(index + 1, min(index + 6, len(grid)))
        )
        if not following:
            continue
        score = _looks_like_header(row) + filled * 0.25
        if score > best_score:
            best_index, best_score = index, score
    return best_index


def _trim(grid):
    """Drop trailing rows and columns that are completely empty."""
    while grid and all(is_blank(c) for c in grid[-1]):
        grid.pop()
    if not grid:
        return grid
    width = 0
    for row in grid:
        for index, cell in enumerate(row):
            if not is_blank(cell):
                width = max(width, index + 1)
    return [list(row[:width]) for row in grid]


def _read_grid(worksheet) -> list[list[Any]]:
    return _trim([list(row) for row in worksheet.iter_rows(values_only=True)])


def build_sheet(worksheet, header_row: int | None = None) -> Sheet:
    """Load a worksheet into a Sheet, detecting the header row if needed."""
    grid = _read_grid(worksheet)
    if not grid:
        return Sheet(name=worksheet.title, header_row=1, columns=[], grid=[])

    index = _find_header_row(grid) if header_row is None else header_row - 1
    index = max(0, min(index, len(grid) - 1))

    labels, seen = [], {}
    for position, cell in enumerate(grid[index]):
        label = cell_text(cell) or f"Column {position + 1}"
        if label in seen:                       # keep duplicate headers distinct
            seen[label] += 1
            label = f"{label} ({seen[label]})"
        else:
            seen[label] = 1
        labels.append(label)

    rows, numbers = [], []
    for offset, raw in enumerate(grid[index + 1:], start=index + 2):
        if all(is_blank(c) for c in raw):
            continue                            # skip blank separator rows
        record = {}
        for position, label in enumerate(labels):
            record[label] = raw[position] if position < len(raw) else None
        rows.append(record)
        numbers.append(offset)

    return Sheet(
        name=worksheet.title,
        header_row=index + 1,
        columns=labels,
        rows=rows,
        row_numbers=numbers,
        grid=grid,
    )


def load_workbook(source, filename: str = "") -> Workbook:
    """Open a workbook read-only and load every sheet the audit needs.

    `source` may be a path or a file-like object (a Streamlit upload).
    The file on disk is never modified.
    """
    book = openpyxl.load_workbook(source, data_only=True, read_only=False)
    try:
        names = list(book.sheetnames)
        resolved, missing = detect_sheets(names)

        sheets: dict[str, Sheet] = {}
        for key, sheet_name in resolved.items():
            # The Condition sheet is normally stored without a header row, so
            # it is loaded raw and interpreted in references.py.
            header = 1 if key == "condition_ref" else None
            sheets[key] = build_sheet(book[sheet_name], header_row=header)
            if key == "condition_ref":
                sheets[key].grid = _read_grid(book[sheet_name])

        return Workbook(
            filename=filename or getattr(source, "name", "") or str(source),
            sheet_names=names,
            sheets=sheets,
            resolved=resolved,
            missing=missing,
        )
    finally:
        book.close()


def sheet_overview(source) -> list[dict]:
    """Light-weight listing of every sheet and its filled row count."""
    book = openpyxl.load_workbook(source, data_only=True, read_only=True)
    try:
        overview = []
        for worksheet in book.worksheets:
            filled = 0
            for row in worksheet.iter_rows(values_only=True):
                if any(not is_blank(c) for c in row):
                    filled += 1
            overview.append({
                "name": worksheet.title,
                "rows": max(0, filled - 1),   # exclude the header row
                "columns": worksheet.max_column or 0,
            })
        return overview
    finally:
        book.close()
