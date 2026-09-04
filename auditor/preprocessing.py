"""
Data cleaning for migration — preprocesses exactly three sheets: Supplier,
Office, and Asset | GPE Information. Every other sheet in the workbook
(Instructions, FAQ, the reference sheets, ...) is left completely untouched.

What each sheet gets:
    Supplier                 remove blank/placeholder Supplier Names,
                              remove duplicate Supplier Names (keep the
                              most complete row)
    Office                   remove duplicate Office Names (keep the most
                              complete row), reformat every date column
    Asset | GPE Information  reformat every date column

"Date column" is any column whose header contains the word "date" —
detected by name, not a fixed list, so a template with an extra date column
is still handled without a code change.

A date is only ever converted when it can be read with real confidence
(see dates.py). Anything else is left exactly as it was and reported
separately — nothing is silently guessed for a migration.
"""

from __future__ import annotations

import datetime as _dt

import openpyxl

from .config import REFERENCE_COLUMNS
from .dates import TARGET_NUMBER_FORMAT, parse_date
from .models import PreprocessResult, SheetCleanupSummary, UnparseableDate
from .utils import cell_text, is_blank, is_placeholder, normalize_for_matching, reduce_name
from .workbook import build_sheet, detect_sheets

PROGRESS_STEPS = [
    "Reading workbook",
    "Cleaning Supplier sheet",
    "Cleaning Office sheet",
    "Cleaning Asset | GPE Information sheet",
    "Preparing the cleaned file",
]


def _find_date_columns(columns: list[str]) -> list[str]:
    return [c for c in columns if "date" in reduce_name(c)]


def _filled_count(row: dict, columns: list[str]) -> int:
    return sum(1 for c in columns if not is_blank(row.get(c)))


def _dedupe_keep_most_complete(rows, key_label: str, columns: list[str]):
    """Group `rows` by normalised `key_label`, and mark every row in a group
    for removal except the one with the fewest blank cells (ties keep the
    first one in sheet order).

    Returns (ids_to_drop, removed_count, kept_value_examples).
    """
    groups: dict[str, list[dict]] = {}
    for _, row in rows:
        key = normalize_for_matching(row.get(key_label))
        groups.setdefault(key, []).append(row)

    drop_ids: set[int] = set()
    examples: list[str] = []
    for group in groups.values():
        if len(group) <= 1:
            continue
        best = max(group, key=lambda r: _filled_count(r, columns))
        for row in group:
            if row is not best:
                drop_ids.add(id(row))
        examples.append(cell_text(best.get(key_label)))
    return drop_ids, len(drop_ids), sorted(set(examples))


def _reformat_dates(rows, date_labels: list[str], day_first: bool, sheet_name: str):
    """Reformat every date column in place on the (row_number, data) tuples."""
    reformatted = 0
    unparseable: list[UnparseableDate] = []
    for row_number, data in rows:
        for label in date_labels:
            raw = data.get(label)
            if raw is None or cell_text(raw) == "":
                continue
            parsed = parse_date(raw, day_first)
            if parsed is None:
                unparseable.append(UnparseableDate(
                    sheet=sheet_name, column=label, row=row_number, value=cell_text(raw),
                ))
                continue
            data[label] = parsed
            reformatted += 1
    return reformatted, unparseable


def _write_rows(worksheet, header_row: int, columns: list[str], rows: list[dict],
                date_columns: set[str]) -> None:
    """Replace every data row below the header with `rows`, in column order."""
    max_row = worksheet.max_row
    if max_row > header_row:
        worksheet.delete_rows(header_row + 1, max_row - header_row)

    for offset, row in enumerate(rows):
        excel_row = header_row + 1 + offset
        for col_index, label in enumerate(columns, start=1):
            value = row.get(label)
            cell = worksheet.cell(row=excel_row, column=col_index, value=value)
            if label in date_columns and isinstance(value, _dt.datetime):
                cell.number_format = TARGET_NUMBER_FORMAT


def _clean_supplier(book, resolved: dict, result: PreprocessResult) -> None:
    sheet_name = resolved.get("supplier_ref")
    if not sheet_name:
        result.missing_sheets.append("Supplier")
        return

    worksheet = book[sheet_name]
    sheet = build_sheet(worksheet)
    summary = SheetCleanupSummary(sheet=sheet_name, key="supplier_ref",
                                  original_rows=len(sheet.rows))

    name_label = sheet.find_column(REFERENCE_COLUMNS["supplier_ref"]["name"])
    if not name_label:
        summary.available = False
        summary.remaining_rows = summary.original_rows
        result.summaries.append(summary)
        return

    rows = list(zip(sheet.row_numbers, sheet.rows))
    kept = [(n, r) for n, r in rows if not (is_blank(r.get(name_label)) or is_placeholder(r.get(name_label)))]
    summary.removed_blank_or_placeholder = len(rows) - len(kept)

    drop_ids, removed, examples = _dedupe_keep_most_complete(kept, name_label, sheet.columns)
    kept = [(n, r) for n, r in kept if id(r) not in drop_ids]
    summary.removed_duplicates = removed
    summary.duplicate_examples = examples
    summary.remaining_rows = len(kept)

    _write_rows(worksheet, sheet.header_row, sheet.columns, [r for _, r in kept], date_columns=set())
    result.summaries.append(summary)


def _clean_office(book, resolved: dict, result: PreprocessResult, day_first: bool) -> None:
    sheet_name = resolved.get("office_ref")
    if not sheet_name:
        result.missing_sheets.append("Office")
        return

    worksheet = book[sheet_name]
    sheet = build_sheet(worksheet)
    summary = SheetCleanupSummary(sheet=sheet_name, key="office_ref",
                                  original_rows=len(sheet.rows))

    date_labels = _find_date_columns(sheet.columns)
    summary.date_columns = date_labels

    rows = list(zip(sheet.row_numbers, sheet.rows))
    name_label = sheet.find_column(REFERENCE_COLUMNS["office_ref"]["name"])
    if name_label:
        # Blank office names are left exactly where they are; only rows with
        # an actual name take part in duplicate detection.
        named = [(n, r) for n, r in rows if not is_blank(r.get(name_label))]
        drop_ids, removed, examples = _dedupe_keep_most_complete(named, name_label, sheet.columns)
        rows = [(n, r) for n, r in rows if id(r) not in drop_ids]
        summary.removed_duplicates = removed
        summary.duplicate_examples = examples

    reformatted, unparseable = _reformat_dates(rows, date_labels, day_first, sheet_name)
    summary.dates_reformatted = reformatted
    summary.dates_unparseable = unparseable
    summary.remaining_rows = len(rows)

    _write_rows(worksheet, sheet.header_row, sheet.columns, [r for _, r in rows],
               date_columns=set(date_labels))
    result.summaries.append(summary)


def _clean_asset_info(book, resolved: dict, result: PreprocessResult, day_first: bool) -> None:
    sheet_name = resolved.get("asset_info")
    if not sheet_name:
        result.missing_sheets.append("Asset | GPE Information")
        return

    worksheet = book[sheet_name]
    sheet = build_sheet(worksheet)
    summary = SheetCleanupSummary(sheet=sheet_name, key="asset_info",
                                  original_rows=len(sheet.rows))

    date_labels = _find_date_columns(sheet.columns)
    summary.date_columns = date_labels

    rows = list(zip(sheet.row_numbers, sheet.rows))
    reformatted, unparseable = _reformat_dates(rows, date_labels, day_first, sheet_name)
    summary.dates_reformatted = reformatted
    summary.dates_unparseable = unparseable
    summary.remaining_rows = len(rows)

    _write_rows(worksheet, sheet.header_row, sheet.columns, [r for _, r in rows],
               date_columns=set(date_labels))
    result.summaries.append(summary)


def clean_workbook(source, day_first: bool, filename: str = "", progress=None):
    """Clean the Supplier, Office and Asset | GPE Information sheets.

    Returns (workbook, PreprocessResult). `workbook` is the openpyxl Workbook
    with the three sheets updated in place — every other sheet is untouched —
    ready to be saved. The file at `source` is never modified; a new workbook
    is built in memory.
    """
    def step(name):
        if progress:
            progress(name)

    step("Reading workbook")
    book = openpyxl.load_workbook(source, data_only=True)
    resolved, _ = detect_sheets(list(book.sheetnames))

    result = PreprocessResult(filename=filename, day_first=day_first)

    step("Cleaning Supplier sheet")
    _clean_supplier(book, resolved, result)

    step("Cleaning Office sheet")
    _clean_office(book, resolved, result, day_first)

    step("Cleaning Asset | GPE Information sheet")
    _clean_asset_info(book, resolved, result, day_first)

    step("Preparing the cleaned file")
    return book, result
